from flask import Flask, render_template, request, redirect, session, send_file, jsonify
import mysql.connector
import io
import zipfile
import os
import uuid
from datetime import datetime
from openpyxl import Workbook

app = Flask(__name__)
app.secret_key = "supersecretkey123"

# -------------------- DATABASE CONFIG --------------------
db_config = {
    'host': '192.168.1.22',
    'port': '3306',
    'user': 'avadh',
    'password': 'Avadh!@#123',
    'database': 'test'
}

MAX_ROWS_PER_FILE = 600000
CHUNK_SIZE = 10000
DOWNLOAD_FOLDER = "/tmp"

# -------------------- PROGRESS STORE --------------------
progress_store = {}

# -------------------- LOGIN --------------------
@app.route('/', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        email = request.form['email'].strip()
        password = request.form['password'].strip()

        try:
            conn = mysql.connector.connect(**db_config)
            cursor = conn.cursor(dictionary=True)
            cursor.execute(
                "SELECT DISTINCT Email, Password, HsCode, PortType FROM Users WHERE Email=%s AND Password=%s",
                (email, password)
            )
            users = cursor.fetchall()
        except Exception as e:
            error = f"Database error: {e}"
            users = []
        finally:
            cursor.close()
            conn.close()

        if users:
            session['users'] = users
            if len(users) == 1:
                session['user'] = users[0]['Email']
                session['port_type'] = users[0]['PortType']
                session['hs_code'] = users[0]['HsCode']
                return redirect('/dashboard')
            else:
                return render_template('choose_port.html', users=users)
        else:
            if not error:
                error = "Invalid Email or Password"

    return render_template('login.html', error=error)

# -------------------- SELECT PORT --------------------
@app.route('/select_port', methods=['POST'])
def select_port():
    index = int(request.form['port_selection'])
    selected_user = session['users'][index]

    session['user'] = selected_user['Email']
    session['port_type'] = selected_user['PortType']
    session['hs_code'] = selected_user['HsCode']

    return redirect('/dashboard')

# -------------------- DASHBOARD --------------------
@app.route('/dashboard', methods=['GET', 'POST'])
def dashboard():
    if 'user' not in session:
        return redirect('/')

    user_hs_code = str(session['hs_code'])
    port_type = session['port_type'].strip().lower().replace(" ", "_")

    table_mapping = {
        "import": "Monthly_import_off_1to31th_Jan26",
        "export": "Monthly_Export_Offline_Jan26",
        "sez_import": "SEZ_I_Off_Jan26",
        "sez_export": "Sez_E_Off_jan26"
    }

    if port_type not in table_mapping:
        return "Invalid PortType"

    table_name = table_mapping[port_type]

    if request.method == 'POST':
        hs_code_input = request.form.get('hs_code', '').strip()

        if hs_code_input and not hs_code_input.startswith(user_hs_code):
            return "Invalid HS Code"

        hs_filter = f"{hs_code_input or user_hs_code}%"

        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # -------------------- COUNT TOTAL ROWS --------------------
        count_query = f"SELECT COUNT(*) as total FROM `{table_name}` WHERE `HS Code` LIKE %s"
        cursor.execute(count_query, (hs_filter,))
        total_rows = cursor.fetchone()['total']

        # -------------------- MAIN QUERY --------------------
        query = f"SELECT * FROM `{table_name}` WHERE `HS Code` LIKE %s"
        cursor.execute(query, (hs_filter,))

        # -------------------- TASK ID --------------------
        task_id = str(uuid.uuid4())
        session['task_id'] = task_id
        progress_store[task_id] = 0

        # -------------------- FILE SETUP --------------------
        filename = f"{user_hs_code}_{port_type}.zip"
        zip_path = os.path.join(DOWNLOAD_FOLDER, filename)
        zip_file = zipfile.ZipFile(zip_path, mode='w', compression=zipfile.ZIP_DEFLATED)

        file_count = 1
        row_count = 0
        processed_rows = 0

        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title="Data")

        columns = [col[0] for col in cursor.description]
        ws.append(columns)

        # -------------------- PROCESS DATA --------------------
        while True:
            rows = cursor.fetchmany(CHUNK_SIZE)
            if not rows:
                break

            for row in rows:
                ws.append([row[col] if row[col] is not None else '' for col in columns])

                row_count += 1
                processed_rows += 1

                # UPDATE PROGRESS
                if total_rows > 0:
                    progress_store[task_id] = int((processed_rows / total_rows) * 100)

                # SPLIT FILE
                if row_count >= MAX_ROWS_PER_FILE:
                    excel_buffer = io.BytesIO()
                    wb.save(excel_buffer)
                    excel_buffer.seek(0)

                    zip_file.writestr(f"data_part_{file_count}.xlsx", excel_buffer.read())

                    file_count += 1
                    row_count = 0

                    wb = Workbook(write_only=True)
                    ws = wb.create_sheet(title="Data")
                    ws.append(columns)

        # -------------------- FINAL FILE --------------------
        if row_count > 0:
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            zip_file.writestr(f"data_part_{file_count}.xlsx", excel_buffer.read())

        zip_file.close()
        cursor.close()
        conn.close()

        # MARK COMPLETE
        progress_store[task_id] = 100

        # LOG DOWNLOAD
        if 'downloads' not in session:
            session['downloads'] = []

        session['downloads'].append({
            'filename': filename,
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })

        return send_file(zip_path, as_attachment=True)

    return render_template(
        'dashboard.html',
        user_port_type=port_type,
        user_hs_code=user_hs_code,
        downloads=session.get('downloads', [])
    )

# -------------------- PROGRESS API --------------------
@app.route('/progress')
def progress():
    task_id = session.get('task_id')
    if not task_id:
        return jsonify({"progress": 0})

    return jsonify({"progress": progress_store.get(task_id, 0)})

# -------------------- LOGOUT --------------------
@app.route('/logout')
def logout():
    session.clear()
    return redirect('/')

# -------------------- RUN --------------------
if __name__ == '__main__':
    app.run(host="0.0.0.0", port=5000, debug=True)
